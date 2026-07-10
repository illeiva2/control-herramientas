"""Importa los datos iniciales desde los dos Excel originales.

Uso:
    python importar_excel.py "CONTROL DE HERRAMIENTAS.xlsm" "Inventario (1).xlsx"

Crea/actualiza data/panol.db. Se puede correr mas de una vez: hace upsert
por codigo/nombre y no duplica movimientos ya importados.
"""
import sys
import sqlite3
from pathlib import Path
from datetime import datetime, date

from openpyxl import load_workbook

BASE = Path(__file__).parent
DB = BASE / "data" / "panol.db"


def norm_codigo(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def norm_fecha(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def ubicacion_texto(modulo, estante):
    m, e = str(modulo or "").strip(), str(estante or "").strip()
    if m.upper() == "PARED":
        return "Pared"
    if e.upper() == "P":
        return f"Puntera modulo {m.rstrip('.0') or m}"
    if m and e:
        return f"Gondola {m} - Estante {e}"
    return m or e or ""


def main(control_path, inventario_path):
    DB.parent.mkdir(exist_ok=True)
    con = sqlite3.connect(DB)
    con.executescript((BASE / "schema.sql").read_text(encoding="utf-8"))

    # ---- Inventario: catalogo de herramientas desde MATRIX ----
    inv = load_workbook(inventario_path, read_only=True, data_only=True)
    ws = inv["MATRIX"]
    n_htas = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        codigo = norm_codigo(row[0])
        if not codigo or row[1] is None:
            continue
        nombre = str(row[1]).strip()
        cantidad = int(row[2]) if isinstance(row[2], (int, float)) else 0
        modulo = norm_codigo(row[3])
        estante = norm_codigo(row[4])
        detalle = str(row[5]).strip() if row[5] else None
        con.execute(
            """INSERT INTO herramientas (codigo, nombre, cantidad, modulo, estante, ubicacion, detalle)
               VALUES (?,?,?,?,?,?,?)
               ON CONFLICT(codigo) DO UPDATE SET
                 nombre=excluded.nombre, cantidad=excluded.cantidad,
                 modulo=excluded.modulo, estante=excluded.estante,
                 ubicacion=excluded.ubicacion, detalle=excluded.detalle""",
            (codigo, nombre, cantidad, modulo, estante, ubicacion_texto(modulo, estante), detalle),
        )
        n_htas += 1

    # ---- Inventario: empleados ----
    ws = inv["Empleados"]
    n_emp = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        con.execute(
            "INSERT INTO empleados (dni, nombre) VALUES (?,?) ON CONFLICT(nombre) DO NOTHING",
            (norm_codigo(row[0]), str(row[1]).strip()),
        )
        n_emp += 1

    # ---- Control: almacenistas, condiciones, movimientos ----
    ctl = load_workbook(control_path, read_only=True, data_only=True)

    for row in ctl["ALMACENISTA"].iter_rows(min_row=2, values_only=True):
        if row[1]:
            con.execute(
                "INSERT INTO almacenistas (dni, nombre) VALUES (?,?) ON CONFLICT(nombre) DO NOTHING",
                (norm_codigo(row[0]), str(row[1]).strip()),
            )

    for row in ctl["CONDICION HTAS"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            con.execute(
                "INSERT INTO condiciones (nombre) VALUES (?) ON CONFLICT(nombre) DO NOTHING",
                (str(row[0]).strip(),),
            )

    def empleado_id(nombre, dni):
        cur = con.execute("SELECT id FROM empleados WHERE nombre = ?", (nombre,))
        r = cur.fetchone()
        if r:
            return r[0]
        return con.execute(
            "INSERT INTO empleados (dni, nombre) VALUES (?,?)", (dni, nombre)
        ).lastrowid

    def almacenista_id(nombre, dni):
        cur = con.execute("SELECT id FROM almacenistas WHERE nombre = ?", (nombre,))
        r = cur.fetchone()
        if r:
            return r[0]
        return con.execute(
            "INSERT INTO almacenistas (dni, nombre) VALUES (?,?)", (dni, nombre)
        ).lastrowid

    def herramienta_id(codigo, nombre):
        cur = con.execute("SELECT id FROM herramientas WHERE codigo = ?", (codigo,))
        r = cur.fetchone()
        if r:
            return r[0]
        return con.execute(
            "INSERT INTO herramientas (codigo, nombre, cantidad) VALUES (?,?,0)",
            (codigo, nombre or codigo),
        ).lastrowid

    n_mov = 0
    # ENTREGA: FECHA DNI PERSONAL CODIGO HERRAMIENTA CANTIDAD DNI2 ALMACENISTA
    for row in ctl["ENTREGA"].iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[3]:
            continue
        fecha = norm_fecha(row[0])
        emp = empleado_id(str(row[2]).strip(), norm_codigo(row[1]))
        hta = herramienta_id(norm_codigo(row[3]), row[4] and str(row[4]).strip())
        cant = int(row[5] or 1)
        alm = almacenista_id(str(row[7]).strip(), norm_codigo(row[6])) if row[7] else None
        dup = con.execute(
            """SELECT 1 FROM movimientos WHERE tipo='ENTREGA' AND fecha=? AND empleado_id=?
               AND herramienta_id=? AND cantidad=?""",
            (fecha, emp, hta, cant),
        ).fetchone()
        if not dup:
            con.execute(
                """INSERT INTO movimientos (tipo, fecha, empleado_id, herramienta_id, cantidad, almacenista_id)
                   VALUES ('ENTREGA',?,?,?,?,?)""",
                (fecha, emp, hta, cant, alm),
            )
            n_mov += 1

    # DEVOLUCION: FECHA DNI PERSONAL CODIGO HERRAMIENTA CANTIDAD CONDICION DNI2 ALMACENISTA
    for row in ctl["DEVOLUCION"].iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[3]:
            continue
        fecha = norm_fecha(row[0])
        emp = empleado_id(str(row[2]).strip(), norm_codigo(row[1]))
        hta = herramienta_id(norm_codigo(row[3]), row[4] and str(row[4]).strip())
        cant = int(row[5] or 1)
        cond = None
        if row[6]:
            con.execute(
                "INSERT INTO condiciones (nombre) VALUES (?) ON CONFLICT(nombre) DO NOTHING",
                (str(row[6]).strip(),),
            )
            cond = con.execute(
                "SELECT id FROM condiciones WHERE nombre = ?", (str(row[6]).strip(),)
            ).fetchone()[0]
        alm = almacenista_id(str(row[8]).strip(), norm_codigo(row[7])) if row[8] else None
        dup = con.execute(
            """SELECT 1 FROM movimientos WHERE tipo='DEVOLUCION' AND fecha=? AND empleado_id=?
               AND herramienta_id=? AND cantidad=?""",
            (fecha, emp, hta, cant),
        ).fetchone()
        if not dup:
            con.execute(
                """INSERT INTO movimientos (tipo, fecha, empleado_id, herramienta_id, cantidad, condicion_id, almacenista_id)
                   VALUES ('DEVOLUCION',?,?,?,?,?,?)""",
                (fecha, emp, hta, cant, cond, alm),
            )
            n_mov += 1

    con.commit()
    tot = {
        t: con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
        for t in ("herramientas", "empleados", "almacenistas", "condiciones", "movimientos")
    }
    con.close()
    print(f"Importadas {n_htas} herramientas del inventario, {n_mov} movimientos nuevos.")
    print("Totales en la base:", tot)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
